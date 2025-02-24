from fastapi import FastAPI, File, UploadFile, HTTPException
from openpyxl import load_workbook
from fastapi.responses import JSONResponse

app = FastAPI()

@app.post("/get-column-data/")
async def get_column_data(file: UploadFile = File(...)):
    # Ensure the file is an Excel file
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx)")

    # Load the source workbook
    try:
        source_workbook = load_workbook(file.file)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error loading workbook: {str(e)}")

    source_sheet_name = "Page 1"
    if source_sheet_name not in source_workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Sheet '{source_sheet_name}' not found in workbook")

    source_sheet = source_workbook[source_sheet_name]

    # Extract all data from column 1
    column_data = []
    for row in source_sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        if row[0] is not None:  # Skip empty cells
            column_data.append(row[0])

    # Return the data as a JSON response
    return JSONResponse(content={"column_data": column_data})

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
